"""Tests for ONS five-year rolling energy efficiency ETL."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from ons_ee_fiveyear_config import EE_ID_HEADERS
from ons_ee_fiveyear_etl import melt_rolling_sheet, read_ee_sheet, transform_workbook


def _write_like_ons(path: Path, sheet_name: str, body: pd.DataFrame) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(sheet_name)
    for _ in range(4):
        ws.append([None])
    ws.append(list(body.columns))
    for _, row in body.iterrows():
        ws.append(row.tolist())
    wb.save(path)


def test_split_metric_period():
    from ons_ee_fiveyear_etl import _split_metric_period

    assert _split_metric_period("All Q2 2008 to Q1 2013") == ("All", "Q2 2008 to Q1 2013")
    assert _split_metric_period("Flats and maisonettes Q2 2020 to Q1 2025") == (
        "Flats and maisonettes",
        "Q2 2020 to Q1 2025",
    )


def test_melt_minimal():
    cols = [
        EE_ID_HEADERS[0],
        EE_ID_HEADERS[1],
        "All Q2 2008 to Q1 2013",
        "New Q2 2008 to Q1 2013",
    ]
    body = pd.DataFrame(
        [["England", "E92000001", 1.5, 2.5]],
        columns=cols,
    )
    tidy = melt_rolling_sheet(body, "1a")
    assert len(tidy) == 2
    assert set(tidy["measure_breakdown"]) == {"All", "New"}
    assert tidy["value"].dtype == pd.Float64Dtype()


def test_read_and_melt_roundtrip(tmp_path: Path):
    cols = [
        EE_ID_HEADERS[0],
        EE_ID_HEADERS[1],
        "All Q2 2008 to Q1 2013",
    ]
    body = pd.DataFrame([["England", "E1", 42.0]], columns=cols)
    path = tmp_path / "t.xlsx"
    _write_like_ons(path, "1a", body)
    raw = read_ee_sheet(path, "1a")
    tidy = melt_rolling_sheet(raw, "1a")
    assert len(tidy) == 1
    assert tidy["value"].iloc[0] == 42.0


def test_transform_workbook_smoke(tmp_path: Path):
    from ons_ee_fiveyear_config import EE_DATA_SHEETS

    cols = [EE_ID_HEADERS[0], EE_ID_HEADERS[1], "All Q2 2008 to Q1 2013"]
    path = tmp_path / "wb.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    for sn in EE_DATA_SHEETS:
        ws = wb.create_sheet(sn)
        for _ in range(4):
            ws.append([None])
        ws.append(list(cols))
        ws.append(["England", "E1", 1.0])
    wb.save(path)

    out = tmp_path / "out"
    transform_workbook(path, out, edition_key="test", write_parquet=True, verbose=False)
    assert (out / "ons_ee_fiveyear_test_1a_tidy.parquet").is_file()
